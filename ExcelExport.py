from SQLWorking import *
import openpyxl as xl
import os.path
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import pandas as pd


def Excel_Export(date):
    date = date
    excelDate = date.replace("/", "-")
    dateBreak = excelDate.split("-")
    temp = pd.Timestamp(dateBreak[2] + "-" + dateBreak[1] + "-" + dateBreak[0])
    day = temp.day_name()

    desktop = os.path.expanduser("~\Desktop\\")
    saveLocation = desktop + "Scehedule For Date " + excelDate + ".xlsx"
    wb = xl.Workbook()
    wb.save(saveLocation)
    wb.create_sheet("Teachers Schedule")
    ws = wb['Teachers Schedule']
    wb.create_sheet("Teachers Free Slots")
    ws = wb['Teachers Free Slots']
    del wb['Sheet']
    ws = wb['Teachers Schedule']
    sheet = wb['Teachers Schedule']
    wb.save(saveLocation)

    data = fetch_values_from_time_table(date)

    thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    center = Alignment(horizontal='center', vertical='center')
    titleColor = PatternFill(start_color='000000',
                          end_color='000000',
                          fill_type='solid')
    headingColor = PatternFill(start_color='000000',
                          end_color='000000',
                          fill_type='solid')
    resultsColor = PatternFill(start_color='FFFFFF',
                          end_color='FFFFFF',
                          fill_type='solid')
    locationColor = PatternFill(start_color='00FFFF',
                          end_color='00FFFF',
                          fill_type='solid')
    deletionColor = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')

    titleFont = Font(color="FFFFFF", size=12)
    headingFont = Font(color="FFFFFF", size=14)

    cell = sheet.cell(1, 1)
    cell.value = "Classes Schedule " + date
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cell.fill = titleColor
    cell.font = titleFont
    cell.alignment = center
    cell = sheet.cell(2, 1)
    cell.value = day + " Class Time Table"
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    cell.fill = titleColor
    cell.font = titleFont
    cell.alignment = center
    cell = sheet.cell(3, 1)
    cell.value = "Time"
    cell.alignment = center
    # cell.font = Font(bold=True)
    cell.font = headingFont
    cell.fill = headingColor
    cell.border = thin_border
    cell = sheet.cell(3, 2)
    cell.value = "Teacher Name"
    cell.alignment = center
    cell.font = Font(bold=True)
    cell.fill = headingColor
    cell.font = headingFont
    cell.border = thin_border
    cell = sheet.cell(3, 3)
    cell.value = "Subject Name"
    cell.alignment = center
    cell.font = headingFont
    cell.fill = headingColor
    cell.border = thin_border
    cell = sheet.cell(3, 4)
    cell.value = "Location"
    cell.alignment = center
    cell.font = headingFont
    cell.fill = headingColor
    cell.border = thin_border
    wb.save(saveLocation)

    row = 4
    for entry in data:
        list = entry.split(" - ")
        time = sheet.cell(row, 1)
        time.value = list[2]
        time.border = thin_border
        subject = sheet.cell(row, 2)
        subject.value = list[1]
        subject.border = thin_border
        teacher = sheet.cell(row, 3)
        teacher.value = list[0]
        teacher.border = thin_border
        room = sheet.cell(row, 4)
        room.value = list[3]
        room.border = thin_border
        row += 1
    wb.save(saveLocation)

    time_slots = []

    for row in range(4, sheet.max_row+1):
        cell = sheet.cell(row, 1)
        if cell.value not in time_slots:
            time_slots.append(cell.value)

    while len(time_slots) > 0:
        for row in range(4, sheet.max_row+1):
            cell = sheet.cell(row, 1)
            if cell.value in time_slots:
                sheet.insert_rows(row)
                time_slots.remove(cell.value)

    slot = 1
    for row in range(1, sheet.max_row+1):
        cell = sheet.cell(row, 1)
        if cell.value is None:
            cell.value = "Slot " + str(slot)
            # sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell.font = headingFont
            cell.fill = headingColor
            cell.alignment = center
            slot += 1

    dim_holder = DimensionHolder(worksheet=ws)

    time_max_length = 0
    subject_name_length = 0
    teacher_name_length = 0
    room_max_length = 0

    for row in range(3, sheet.max_row + 1):
        time = sheet.cell(row, 1)
        subject = sheet.cell(row, 2)
        teacher = sheet.cell(row, 3)
        room = sheet.cell(row, 4)
        try:
            if len(time.value) > time_max_length:
                time_max_length = len(time.value)
            if len(subject.value) > subject_name_length:
                subject_name_length = len(subject.value)
            if len(teacher.value) > teacher_name_length:
                teacher_name_length = len(teacher.value)
            if len(room.value) > room_max_length:
                room_max_length = len(room.value)
        except TypeError:
            pass


    dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=time_max_length + 3)
    dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=subject_name_length + 3)
    dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=3, max=3, width=teacher_name_length + 3)
    dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=4, max=4, width=room_max_length + 3)

    ws.column_dimensions = dim_holder

    for row in range(3, sheet.max_row + 1):
        slot1 = sheet.cell(row, 1)
        if slot1.value == "Slot 1":
            sheet.delete_rows(row)

    for row in range(3, sheet.max_row + 1):
        slot1 = sheet.cell(row, 1)
        try:
            if "Slot" in slot1.value:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        except TypeError:
            pass

    wb.save(saveLocation)

    teachers_in_time_table = []
    teachers_with_time_in_time_table = []
    teachers_free = []

    All_teachers = fetch_values_from_teachers_table()
    teachers_in_final_table = fetch_values_from_time_table(date)
    all_time_slots = fetch_values_from_time_slots_table()



    for teacher in teachers_in_final_table:
        list = teacher.split(" - ")
        teachers_in_time_table.append(list[0])
        teachers_with_time_in_time_table.append(list[0] + " - " + list[2])
    for teacher in All_teachers:
        list = teacher.split(" - ")
        tutor = list[0]
        if tutor not in teachers_in_time_table:
            for time in all_time_slots:
                caption = tutor + " - " + time
                if caption not in teachers_free:
                    teachers_free.append(caption)
        elif tutor in teachers_in_time_table:
            for time in all_time_slots:
                if str(tutor + " - " + time) in teachers_with_time_in_time_table:
                    pass
                else:
                    caption = tutor + " - " + time
                    if caption not in teachers_free:
                        teachers_free.append(caption)

    ws = wb['Teachers Free Slots']
    sheet = wb['Teachers Free Slots']

    cell = sheet.cell(1, 1)
    cell.value = day + " Teacher Free Time Slots"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    cell.fill = titleColor
    cell.font = titleFont
    cell.alignment = center
    cell = sheet.cell(2, 1)
    cell.value = "Teacher Name"
    cell.alignment = center
    # cell.font = Font(bold=True)
    cell.font = headingFont
    cell.fill = headingColor
    cell.border = thin_border
    cell = sheet.cell(2, 2)
    cell.value = "Teacher Time"
    cell.alignment = center
    cell.font = Font(bold=True)
    cell.fill = headingColor
    cell.font = headingFont
    cell.border = thin_border

    sort_with_time = []
    for items in teachers_free:
        list = items.split(" - ")
        sort_with_time.append(list[1] + " - " + list[0])
    sort_with_time.sort()


    row = 3
    for items in sort_with_time:
        list = items.split(" - ")
        teacher = sheet.cell(row, 1)
        teacher.value = list[1]
        teacher.border = thin_border
        time = sheet.cell(row, 2)
        time.value = list[0]
        time.border = thin_border
        row += 1

    wb.save(saveLocation)

    time_slots = []

    for row in range(3, sheet.max_row+1):
        cell = sheet.cell(row, 2)
        if cell.value not in time_slots:
            time_slots.append(cell.value)

    time_slots_reserved = time_slots.copy()

    while len(time_slots) > 0:
        for row in range(3, sheet.max_row+1):
            cell = sheet.cell(row, 2)
            if cell.value in time_slots:
                sheet.insert_rows(row)
                time_slots.remove(cell.value)

    for row in range(3, sheet.max_row+1):
        cell = sheet.cell(row, 1)
        if cell.value is None:
            cell.value = time_slots_reserved[0]
            time_slots_reserved.remove(cell.value)
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell.alignment = center
            cell.fill = headingColor
            cell.font = headingFont
            cell.border = thin_border

    teacher_name_length = 0
    time_max_length = 0

    for row in range(2, sheet.max_row + 1):
        teacher = sheet.cell(row, 1)
        time = sheet.cell(row, 2)
        try:
            if len(teacher.value) > teacher_name_length:
                teacher_name_length = len(teacher.value)
            if len(time.value) > time_max_length:
                time_max_length = len(time.value)

        except TypeError:
            pass

    dim_holder = DimensionHolder(worksheet=ws)

    dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=teacher_name_length + 3)
    dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=time_max_length + 3)

    ws.column_dimensions = dim_holder

    wb.save(saveLocation)
    os.startfile(saveLocation)

