import openpyxl as xl
from SQLWorking import *
from openpyxl.styles import Font
import os

def OpenExcelFile():
    locationExcel = 'import.xlsx'
    wb = xl.load_workbook(locationExcel, data_only=True)
    # wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
    sheet = wb.worksheets[0]

    sheet.delete_rows(3, sheet.max_row+1)
    wb.save('import.xlsx')

    os.startfile("import.xlsx")



def SaveToSQL():
    locationExcel = 'import.xlsx'
    wb = xl.load_workbook(locationExcel, data_only=True)
    #wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
    sheet = wb.worksheets[0]

    for row in range(3, sheet.max_row + 1):
        room = sheet.cell(row, 1)
        if room.value is not None:
            entry_in_rooms_table(room.value)

        time_slot = sheet.cell(row, 2)
        if time_slot.value is not None:
            entry_in_time_slots_table(time_slot.value)

        teacher_name = sheet.cell(row, 3)
        if teacher_name.value is not None:
            name = teacher_name.value
        else:
            name = ""
        teacher_subject = sheet.cell(row, 4)
        if teacher_subject.value is not None:
            subject = teacher_subject.value
        else:
            subject = ""
        teacher_time = sheet.cell(row, 5)
        if teacher_time.value is not None:
            time = teacher_time.value
        else:
            time = ""
        teacher_room = sheet.cell(row, 6)
        if teacher_room.value is not None:
            room = teacher_room.value
        else:
            room = ""
        entry_in_teachers_table(name, subject, time, room)


