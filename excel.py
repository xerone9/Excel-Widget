import openpyxl as xl

columns = {
        "YouTube": "3",
        "Instagram": "4",
        "Reel": "5",
        "Photo": "6",
        }


def excel_entry(date, recepie, category, amount):
    selections_made = []

    for selection, rate in zip(category, amount):
        selections_made.append(selection + " = " + rate)

    try:
        locationExcel = 'data.xlsx'
        wb = xl.load_workbook(locationExcel, data_only=True)
        sheet = wb.worksheets[0]
        next_row = 1
        for row in range(1, sheet.max_row + 1):
            next_row += 1
        datex = sheet.cell(next_row, 1)
        datex.value = date
        recepiex = sheet.cell(next_row, 2)
        recepiex.value = recepie
        for items in selections_made:
            value = str(items).split(" = ")
            cell = sheet.cell(next_row, int(columns.get(value[0], value[0])))
            cell.value = int(value[1])
        totalx = 0
        for col in range(3, sheet.max_column):
            cell = sheet.cell(next_row, col)
            if cell.value is not None:
                totalx += int(cell.value)
        total = sheet.cell(next_row, 7)
        total.value = totalx
        wb.save(locationExcel)
    except PermissionError:
        return "Excel Open"